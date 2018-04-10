#-*- coding: utf-8 -*-
import DiscoverVulByDate
from SysSpider import *
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import *
import threading
import GlobalVar
import Queue
import time
import json
import sys
import os

def gather_cve_id(vuldata):
    #skip the low vul
    cveid = []
    for x in vuldata:
        if x[1] == u'低':
            continue
        if x[2] != '':
            cveid.append(x[2])
    return cveid

class WebDriverControl:
    def __init__(self, waits, driver_path = ''):
        self.waits = waits
        self.driver_path = driver_path

    def start_driver(self):
        capa = DesiredCapabilities.CHROME
        capa["pageLoadStrategy"] = "none"
        capa['loggingPrefs'] = {'performance': 'ALL'}
        if self.driver_path == '':
            #empty means the driver binary in current directory
            self.browser = webdriver.Chrome(desired_capabilities=capa)
        else:
            self.browser = webdriver.Chrome(executable_path=self.driver_path, desired_capabilities=capa)

    def stop_driver(self):
        self.browser.quit()

    def getHttpStatus(self):
        for responseReceived in self.browser.get_log('performance'):
            try:
                response = json.loads(responseReceived[u'message'])[u'message'][u'params'][u'response']
                if response[u'url'] == self.browser.current_url:
                    return (response[u'status'], response[u'statusText'])
            except:
                pass
        return None

    def implicit_get_page_source(self, url):
        #the driver cannot wait the page load compelete as the setting
        #self.browser.implicitly_wait(100)
        self.browser.get(url)
        time.sleep(5)
        return self.browser.page_source

    def explicit_get_page_source(self, url, locator):
        self.browser.get(url)
        try:
            WebDriverWait(self.browser, self.waits, 0.5).until(EC.presence_of_element_located(locator))
            return self.browser.page_source
        except TimeoutException:
            #print 'Wait page load timeout'
            return ''
        finally:
            pass

class WoodpeckerControl:
    def __init__(self, cveid, data_path, driver, waits, tnum):
        self.data_path = data_path
        self.driver_path = driver
        self.wait = waits
        self.tnum = tnum
        self.suse_key = []
        self.redhat_key = []
        for item in GlobalVar.sys:
            if item.find('SUSE') >= 0:
                for item2 in GlobalVar.sys[item]:
                    self.suse_key.append(item2)
            if item.find('REDHAT') >= 0:
                for item2 in GlobalVar.sys[item]:
                    self.redhat_key.append(item2)
        self.cveid = cveid
        self.tqueue = Queue.Queue(10000)
        self.rqueue = Queue.Queue(10000)
        self.threads = []

    def write_exp(self, vuldata):
        if not os.path.exists(self.data_path):
            wb_ini = Workbook()
            ws = wb_ini.active
            ws['A1'] = u'漏洞名称'
            ws['B1'] = u'危险等级'
            ws['C1'] = u'CVE编号'
            ws['D1'] = u'漏洞描述'
            ws['E1'] = u'修复建议'
            ws['F1'] = u'发现时间'
            ws['G1'] = u'操作系统'
            ws['H1'] = u'版本信息'
            wb_ini.save(self.data_path)
        wb = load_workbook(self.data_path)
        #name = wb.get_sheet_names()[0]
        name = wb.sheetnames[0]
        #print name
        sheet1 = wb[name]
        row_num = len(tuple(sheet1.rows)) + 1
        while not self.rqueue.empty():
            d = self.rqueue.get(block=True, timeout=10)
            sheet1.cell(row_num, 1).value = d.keys()[0]
            if len(d[d.keys()[0]]) == 0:
                sheet1.cell(row_num, 2).value = 'No info found'
                row_num += 1
                continue
            for index in d[d.keys()[0]]:
                for x in vuldata:
                    if x[2] == d.keys()[0]:
                        sdata = x
                        break
                col = 1
                for y in sdata:
                    sheet1.cell(row_num, col).value = y
                    col += 1
                sheet1.cell(row_num, col).value = index[0]
                col += 1
                sheet1.cell(row_num, col).value = index[1]
                row_num += 1
        wb.save(self.data_path)

    def main_control(self, vuldata):
        self.task_load()
        self.start_woodpecker_thread()
        while True:
            if not self.rqueue.empty():
                self.write_exp(vuldata)
            else:
                if self.tqueue.empty():
                    alive = False
                    for t in self.threads:
                        if t.is_alive():
                            alive = True
                            break
                    if alive:
                        continue
                    else:
                        break
                else:
                    continue
            time.sleep(2)

    def task_load(self):
        for id in self.cveid:
            self.tqueue.put(GlobalVar.suse_prefix + id, block=True, timeout=10)
            self.tqueue.put(GlobalVar.redhat_prefix + id, block=True, timeout=10)

    def start_woodpecker_thread(self):
        for x in xrange(self.tnum):
            self.threads.append(threading.Thread(target = self.woodpecker_thread))
        print GlobalVar.time_stamp() + '%d web spider engines have been created' % self.tnum
        for t in self.threads:
            t.setDaemon(True)
            t.start()

    def woodpecker_thread(self):
        driver = WebDriverControl(self.wait, self.driver_path)
        driver.start_driver()
        while not self.tqueue.empty():
            task = self.tqueue.get(block=True, timeout=20)
            print '\rTask: %s are excuted' % task ,
            sys.stdout.flush()
            if task.find('redhat') >= 0:
                redhat_a = redhat_analyze(self.redhat_key, task, driver)
                data = redhat_a.start_analyze()
            if task.find('suse') >= 0:
                suse_a = suse_analyze(self.suse_key, task, driver)
                data = suse_a.start_analyze()
            while True:
                try:
                    self.rqueue.put(data, block=True, timeout=20)
                    break
                except Exception, e:
                    continue


def load_config():
    ip = ''
    account = ''
    date_range = []
    driver_path = ''
    vul_path = ''
    spider_path = ''
    waits = 0
    tnum = 0

    for line in open('config').readlines():
        if line.find('spider_data') >= 0:
            spider_path = line.split('=')[1].strip()
        if line.find('web_driver') >= 0:
            driver_path = line.split('=')[1].strip()
        if line.find('page_load_timeout') >= 0:
            waits = int(line.split('=')[1].strip())
        if line.find('threads_num') >= 0:
            tnum = int(line.split('=')[1].strip())
        if line.find('scanner_ip') >= 0:
            ip = line.split('=')[1].strip()
        if line.find('account') >= 0:
            account = line.split('=')[1].strip()
        if line.find('date_range') >= 0:
            tmp = line.split('=')[1].strip()
            date_range = [tmp.split(',')[0], tmp.split(',')[1]]
        if line.find('vul_path') >= 0:
            vul_path = line.split('=')[1].strip()
    return (ip, account, date_range, driver_path, vul_path, spider_path, waits, tnum)

if __name__ == '__main__':
    (ip, account, date_range, driver, vul_path, spider_path, timeout, tnum) = load_config()
    engine = DiscoverVulByDate.DiscoverVul(ip, account , date_range, driver, vul_path)
    vuldata = engine.discover_start()
    engine.save_vul_data()
    cves = gather_cve_id(vuldata)
    GlobalVar.sys_keyword_install()
    #cves = [u'CVE-2017-6014', u'CVE-2017-2996', u'CVE-2017-2993', u'CVE-2017-2992', u'CVE-2017-2991', u'CVE-2017-2990', u'CVE-2017-2988', u'CVE-2017-2987', u'CVE-2017-2986', u'CVE-2017-2985', u'CVE-2017-2982', u'CVE-2017-0037']
    #cves= [u'CVE-2015-5346']
    #cves= [u'CVE-2018-1058']
    control = WoodpeckerControl(cves, spider_path, driver, timeout, tnum)
    control.main_control(vuldata)
    print '\n' + GlobalVar.time_stamp() + 'Spider data has been saved in %s' % spider_path
    print 'All Done!'
    #vuldata.save_vul_data()
